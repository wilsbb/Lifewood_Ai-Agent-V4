import os
import json
import base64
import requests as http_requests
from datetime import datetime, timedelta
from decimal import Decimal

from django.conf import settings
from django.db.models import Sum, Count, Avg, Min, Max, Q
from django.db.models.functions import TruncMonth
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.utils import timezone

from .models import Receipt, Conversation, ChatMessage


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def require_auth(func):
    def wrapper(request, *args, **kwargs):
        if not request.user or not request.user.is_authenticated:
            return JsonResponse({'error': 'Not authenticated'}, status=401)
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


def _is_n8n_request(request):
    agent_secret = os.environ.get('N8N_AGENT_SECRET', '')
    request_secret = request.headers.get('X-Agent-Secret', '')
    return bool(agent_secret and request_secret == agent_secret)


def _get_n8n_credentials():
    from google_drive.models import GoogleDriveToken
    from google_drive.utils import get_credentials_from_token
    token = GoogleDriveToken.objects.first()
    if not token:
        return None
    return get_credentials_from_token(token)


def parse_date_range(request):
    today = timezone.now().date()
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    try:
        start = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else today.replace(day=1)
        end = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else today
    except ValueError:
        start = today.replace(day=1)
        end = today
    return start, end


def get_user_receipts(user):
    """Returns receipts belonging to this user OR receipts with no user assigned."""
    return Receipt.objects.filter(Q(user=user) | Q(user__isnull=True))


# ─────────────────────────────────────────────
# CHAT — INTERNAL AI (no n8n needed)
# ─────────────────────────────────────────────

def _build_analytics_context(user):
    """
    Fetch analytics from the DB for the AI system prompt.
    Includes ALL-TIME folder-based totals (the user's true categorization)
    plus this month's summary and recent receipts.
    """
    today = timezone.now().date()
    start = today.replace(day=1)
    all_receipts = get_user_receipts(user).filter(status='processed')

    # ── This month ─────────────────────────────────────────────────────────
    base_qs = all_receipts.filter(expense_date__range=[start, today])

    prev_end = start - timedelta(days=1)
    prev_start = prev_end.replace(day=1)
    prev_total = all_receipts.filter(
        expense_date__range=[prev_start, prev_end],
    ).aggregate(t=Sum('total'))['t'] or Decimal('0')

    summary = base_qs.aggregate(
        total_spend=Sum('total'),
        total_vat=Sum('vat_amount'),
        transaction_count=Count('id'),
        avg_transaction=Avg('total'),
    )

    # ── ALL-TIME spending by folder (user-defined category) ────────────────
    # drive_folder_name is the authoritative category — set by the user
    # when they physically organised receipts into Google Drive folders.
    by_folder_alltime = list(
        all_receipts
        .exclude(drive_folder_name='')
        .values('drive_folder_name')
        .annotate(total=Sum('total'), vat=Sum('vat_amount'), count=Count('id'))
        .order_by('-total')
    )

    # ── This month spending by folder ─────────────────────────────────────
    by_folder_month = list(
        base_qs
        .exclude(drive_folder_name='')
        .values('drive_folder_name')
        .annotate(total=Sum('total'), vat=Sum('vat_amount'), count=Count('id'))
        .order_by('-total')
    )

    # ── Recent receipts (last 15, all time) ───────────────────────────────
    recent = list(
        all_receipts
        .order_by('-expense_date')[:15]
        .values('business_name', 'drive_folder_name', 'expense_category',
                'document_type', 'total', 'vat_amount', 'expense_date',
                'description', 'tin', 'receipt_number', 'vat_type', 'bir_permit_number')
    )

    total_spend = summary['total_spend'] or Decimal('0')
    total_vat   = summary['total_vat']   or Decimal('0')
    tx_count    = summary['transaction_count'] or 0
    avg_tx      = summary['avg_transaction'] or Decimal('0')

    change_pct = 0.0
    if prev_total > 0:
        change_pct = float((total_spend - prev_total) / prev_total * 100)

    folder_alltime_lines = '\n'.join(
        f"  - {f['drive_folder_name']}: PHP {f['total']} "
        f"(VAT: PHP {f['vat']}, {f['count']} receipts)"
        for f in by_folder_alltime
    ) or '  (no folder data yet)'

    folder_month_lines = '\n'.join(
        f"  - {f['drive_folder_name']}: PHP {f['total']} ({f['count']} receipts)"
        for f in by_folder_month
    ) or '  (no receipts this month)'

    recent_lines = '\n'.join(
        f"  - {r['expense_date'] or 'Unknown'} | "
        f"Folder: {r['drive_folder_name'] or 'Uncategorized'} | "
        f"{r['business_name'] or 'Unknown'} | PHP {r['total']} | "
        f"VAT: {r['vat_type']} | TIN: {r['tin'] or 'missing'} | "
        f"Receipt#: {r['receipt_number'] or 'missing'}"
        for r in recent
    ) or '  (no receipts yet)'

    return f"""=== THIS MONTH SUMMARY ({start} to {today}) ===
Total Spent      : PHP {total_spend}
VAT Paid         : PHP {total_vat}
Transactions     : {tx_count}
Avg per Receipt  : PHP {avg_tx}
vs Last Month    : {change_pct:+.1f}% (last month: PHP {prev_total})

=== THIS MONTH BY FOLDER (user-defined categories) ===
{folder_month_lines}

=== ALL-TIME SPENDING BY FOLDER (user-defined categories) ===
{folder_alltime_lines}

=== RECENT RECEIPTS (last 15) ===
{recent_lines}"""


def _build_memory_context(user, query, limit=6):
    """Fetch relevant past messages for context."""
    qs = (
        ChatMessage.objects
        .filter(conversation__user=user)
        .select_related('conversation')
        .order_by('-created_at')
    )
    if query:
        qs = qs.filter(content__icontains=query)
    messages = list(qs[:limit])
    if not messages:
        return ''
    lines = '\n'.join(f"[{m.role}]: {m.content[:300]}" for m in messages)
    return f"=== RELEVANT PAST CONVERSATIONS ===\n{lines}"


def _call_openrouter(messages):
    """Call OpenRouter (OpenAI-compatible) and return the reply text + usage."""
    api_key = os.environ.get('OPENROUTER_API_KEY', '')
    if not api_key:
        raise ValueError('OPENROUTER_API_KEY is not set')

    response = http_requests.post(
        'https://openrouter.ai/api/v1/chat/completions',
        headers={
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
            'HTTP-Referer': os.environ.get('FRONTEND_URL', 'https://lifewood.ai'),
            'X-Title': 'Lifewood Expense AI',
        },
        json={
            'model': 'openai/gpt-4o',
            'max_tokens': 1500,
            'messages': messages,
        },
        timeout=60,
    )
    response.raise_for_status()
    data = response.json()
    reply = data['choices'][0]['message']['content']
    usage = data.get('usage', {})
    return reply, usage


@csrf_exempt
@require_POST
@require_auth
def send_message(request):
    """
    Handles chat entirely within Django — no n8n dependency.
    1. Loads conversation history
    2. Fetches analytics + memory from DB
    3. Calls OpenRouter/GPT-4o directly
    4. Saves and returns the reply
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_message = body.get('message', '').strip()
    conversation_id = body.get('conversation_id')

    if not user_message:
        return JsonResponse({'error': 'Message cannot be empty'}, status=400)


    # ── Export intent detection ────────────────────────────────────────────
    # Detect BEFORE normal chat processing so we short-circuit cleanly.
    EXPORT_KEYWORDS = ['export', 'excel', 'xlsx', 'spreadsheet', 'download receipts']
    msg_lower = user_message.lower()
    if any(kw in msg_lower for kw in EXPORT_KEYWORDS):

        # ── Get all folder names from DB for matching ──────────────────────
        all_folder_names = list(
            get_user_receipts(request.user)
            .filter(status='processed')
            .exclude(drive_folder_name='')
            .values_list('drive_folder_name', flat=True)
            .distinct()
        )
        folder_list_text = '\n'.join(f'  - {f}' for f in all_folder_names) or '  (none)'

        # ── Ask GPT to extract the folder name the user is referring to ────
        folder_filter = ''
        try:
            api_key = os.environ.get('OPENROUTER_API_KEY', '')
            if api_key and all_folder_names:
                extract_resp = http_requests.post(
                    'https://openrouter.ai/api/v1/chat/completions',
                    headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
                    json={
                        'model': 'openai/gpt-4o',
                        'max_tokens': 80,
                        'messages': [{
                            'role': 'user',
                            'content': (
                                f'The user said: "{user_message}"\n\n'
                                f'Available folder names:\n{folder_list_text}\n\n'
                                'Which folder name from the list above is the user referring to?\n'
                                'Return ONLY the exact folder name from the list, or "ALL" if they want all folders, '
                                'or "UNKNOWN" if unclear. No explanation, just the name.'
                            )
                        }],
                    },
                    timeout=15,
                )
                extracted = extract_resp.json()['choices'][0]['message']['content'].strip().strip('"')
                if extracted not in ('ALL', 'UNKNOWN', '') and extracted in all_folder_names:
                    folder_filter = extracted
                elif extracted not in ('ALL', 'UNKNOWN', ''):
                    # Fuzzy fallback: find closest match
                    ext_lower = extracted.lower()
                    for fname in all_folder_names:
                        if ext_lower in fname.lower() or fname.lower() in ext_lower:
                            folder_filter = fname
                            break
        except Exception as e:
            print(f'Export folder extraction error: {e}')
            # Fall back to simple substring match
            for fname in all_folder_names:
                if fname.lower() in msg_lower:
                    folder_filter = fname
                    break

        # ── Build download URL ─────────────────────────────────────────────
        from urllib.parse import urlencode, quote
        base_url     = os.environ.get('BACKEND_URL', '').rstrip('/')
        query_string = f'?folder={quote(folder_filter)}' if folder_filter else ''
        download_url = f'{base_url}/api/billing/receipts/export/{query_string}'

        # Count matching receipts
        qs_count = get_user_receipts(request.user).filter(status='processed')
        if folder_filter:
            qs_count = qs_count.filter(drive_folder_name__icontains=folder_filter)
        receipt_count = qs_count.count()

        if folder_filter:
            reply = (
                f"Your Excel export for the **{folder_filter}** folder is ready — "
                f"{receipt_count} receipt{'s' if receipt_count != 1 else ''}. "
                f"Click the download button below."
            )
        elif receipt_count > 0:
            reply = (
                f"Your Excel export of all receipts is ready — "
                f"{receipt_count} total receipt{'s' if receipt_count != 1 else ''}. "
                f"Click the download button below."
            )
        else:
            reply = (
                "I couldn't find any matching receipts to export. "
                "Please check the folder name and try again."
            )

        # ── Save to conversation ───────────────────────────────────────────
        if conversation_id:
            try:
                conversation = Conversation.objects.get(id=conversation_id, user=request.user)
            except Conversation.DoesNotExist:
                conversation = Conversation.objects.create(user=request.user, title=user_message[:60])
        else:
            conversation = Conversation.objects.create(user=request.user, title=user_message[:60])

        user_chat_msg = ChatMessage.objects.create(
            conversation=conversation, role='user', content=user_message)
        agent_meta = {
            'download_url':   download_url,
            'export':         True,
            'folder_filter':  folder_filter,
            'receipt_count':  receipt_count,
        }
        agent_msg = ChatMessage.objects.create(
            conversation=conversation, role='agent',
            content=reply, metadata=agent_meta,
        )
        conversation.save()

        return JsonResponse({
            'conversation_id':  conversation.id,
            'user_message_id':  user_chat_msg.id,
            'agent_message_id': agent_msg.id,
            'reply':            reply,
            'metadata':         agent_meta,
        })


    # ── Get or create conversation ─────────────────────────────────────────
    if conversation_id:
        try:
            conversation = Conversation.objects.get(id=conversation_id, user=request.user)
        except Conversation.DoesNotExist:
            return JsonResponse({'error': 'Conversation not found'}, status=404)
    else:
        title = user_message[:60] + ('...' if len(user_message) > 60 else '')
        conversation = Conversation.objects.create(user=request.user, title=title)

    # ── Save user message ──────────────────────────────────────────────────
    user_chat_msg = ChatMessage.objects.create(
        conversation=conversation,
        role='user',
        content=user_message,
    )

    # ── Build conversation history (last 10 turns) ─────────────────────────
    history_qs = list(
        conversation.messages
        .exclude(id=user_chat_msg.id)
        .order_by('-created_at')[:10]
    )
    history_messages = [
        {
            'role': 'assistant' if m.role == 'agent' else 'user',
            'content': m.content,
        }
        for m in reversed(history_qs)
    ]

    # ── Build system prompt with live data ────────────────────────────────
    try:
        analytics_context = _build_analytics_context(request.user)
    except Exception as e:
        print(f'Analytics context error: {e}')
        analytics_context = '(analytics unavailable)'

    try:
        memory_context = _build_memory_context(request.user, user_message)
    except Exception as e:
        print(f'Memory context error: {e}')
        memory_context = ''

    system_prompt = f"""You are Lifewood's AI finance assistant. You help users understand their expense history, receipts, and BIR (Bureau of Internal Revenue) compliance in the Philippines.

Guidelines:
- Be concise, friendly, and professional
- Always reference the actual data provided below when answering
- Format all currency as PHP X,XXX.XX
- If asked about something not in the data, say so honestly
- You can only see this user's own expense data
- When asked about BIR compliance, reference Philippine BIR rules (VAT registration, TIN requirements, official receipts, BIR permit numbers)
- Flag any receipts missing TIN, receipt number, or BIR permit number as compliance risks

IMPORTANT — HOW CATEGORIES WORK:
The user organises receipts by placing them into named Google Drive folders. The folder name is the TRUE and AUTHORITATIVE expense category. For example, a folder called "Condo Dues" contains all condo-related receipts, "Admin Expenses" contains all admin receipts, etc.
When the user asks "how much did we spend on X?", match X against the folder names in the data below (case-insensitive, partial match is fine). Always answer using folder-based totals, not the system-guessed expense_category field.
If the user asks about a folder that has no data, say so clearly.

{analytics_context}

{memory_context}""".strip()

    # ── Call OpenRouter ────────────────────────────────────────────────────
    messages = [
        {'role': 'system', 'content': system_prompt},
        *history_messages,
        {'role': 'user', 'content': user_message},
    ]

    try:
        reply, usage = _call_openrouter(messages)
        agent_metadata = {
            'model': 'openai/gpt-4o',
            'input_tokens': usage.get('prompt_tokens'),
            'output_tokens': usage.get('completion_tokens'),
            'total_tokens': usage.get('total_tokens'),
        }
    except Exception as e:
        print(f'OpenRouter error: {e}')
        reply = 'I encountered an issue reaching the AI. Please try again in a moment.'
        agent_metadata = {}

    # ── Save agent reply ───────────────────────────────────────────────────
    agent_chat_msg = ChatMessage.objects.create(
        conversation=conversation,
        role='agent',
        content=reply,
        metadata=agent_metadata,
    )
    conversation.save()

    return JsonResponse({
        'conversation_id': conversation.id,
        'user_message_id': user_chat_msg.id,
        'agent_message_id': agent_chat_msg.id,
        'reply': reply,
        'metadata': agent_metadata,
    })


@require_GET
@require_auth
def get_conversation_history(request):
    conversation_id = request.GET.get('conversation_id')
    if not conversation_id:
        return JsonResponse({'error': 'conversation_id is required'}, status=400)
    try:
        conversation = Conversation.objects.get(id=conversation_id, user=request.user)
    except Conversation.DoesNotExist:
        return JsonResponse({'error': 'Conversation not found'}, status=404)

    messages = conversation.messages.all().values('id', 'role', 'content', 'metadata', 'created_at')
    return JsonResponse({
        'conversation_id': conversation.id,
        'title': conversation.title,
        'messages': [{**msg, 'created_at': msg['created_at'].isoformat()} for msg in messages],
    })


@require_GET
@require_auth
def list_conversations(request):
    conversations = (
        Conversation.objects
        .filter(user=request.user)
        .annotate(message_count=Count('messages'))
        .values('id', 'title', 'created_at', 'updated_at', 'message_count')
        .order_by('-updated_at')
    )
    return JsonResponse({
        'conversations': [
            {**conv, 'created_at': conv['created_at'].isoformat(), 'updated_at': conv['updated_at'].isoformat()}
            for conv in conversations
        ]
    })


# ─────────────────────────────────────────────
# MEMORY ENDPOINT (kept for n8n compatibility)
# ─────────────────────────────────────────────

@csrf_exempt
@require_GET
def chat_memory(request):
    if _is_n8n_request(request):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user_id = request.GET.get('user_id')
        if not user_id:
            return JsonResponse({'error': 'user_id is required for agent requests'}, status=400)
        user = User.objects.filter(id=user_id).first()
        if not user:
            return JsonResponse({'error': 'User not found'}, status=404)
    elif request.user and request.user.is_authenticated:
        user = request.user
    else:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    query = request.GET.get('query', '').strip()
    limit = min(int(request.GET.get('limit', 8)), 20)

    base_qs = (
        ChatMessage.objects
        .filter(conversation__user=user)
        .select_related('conversation')
        .order_by('-created_at')
    )
    if query:
        base_qs = base_qs.filter(Q(content__icontains=query))

    messages_page = base_qs[:limit]
    return JsonResponse({
        'user_id': user.id,
        'query': query,
        'memories': [
            {
                'id': m.id,
                'role': m.role,
                'content': m.content[:600] + ('...' if len(m.content) > 600 else ''),
                'conversation_id': m.conversation_id,
                'conversation_title': m.conversation.title,
                'created_at': m.created_at.isoformat(),
            }
            for m in messages_page
        ],
        'total_found': base_qs.count(),
    })


# ─────────────────────────────────────────────
# RECEIPT ENDPOINTS
# ─────────────────────────────────────────────

@csrf_exempt
@require_POST
def save_receipt(request):
    if not _is_n8n_request(request):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    drive_file_id = body.get('drive_file_id')
    if not drive_file_id:
        return JsonResponse({'error': 'drive_file_id is required'}, status=400)

    from django.contrib.auth import get_user_model
    User = get_user_model()
    user = None
    user_id = body.get('user_id')
    if user_id:
        user = User.objects.filter(id=user_id).first()

    expense_date = None
    date_str = body.get('expense_date')
    if date_str:
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                expense_date = datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue

    receipt, created = Receipt.objects.update_or_create(
        drive_file_id=drive_file_id,
        defaults={
            'user': user,
            'drive_file_name': body.get('drive_file_name', ''),
            'drive_folder_id': body.get('drive_folder_id', ''),
            'drive_folder_name': body.get('drive_folder_name', ''),
            'status': body.get('status', 'processed'),
            'ocr_raw_text': body.get('ocr_raw_text', ''),
            'ocr_processed_at': timezone.now(),
            'document_type': body.get('document_type', 'unknown'),
            'vat_type': body.get('vat_type', 'unknown'),
            'expense_category': body.get('expense_category', 'uncategorized'),
            'business_name': body.get('business_name', ''),
            'business_address': body.get('business_address', ''),
            'tin': body.get('tin', ''),
            'receipt_number': body.get('receipt_number', ''),
            'bir_permit_number': body.get('bir_permit_number', ''),
            'expense_date': expense_date,
            'description': body.get('description', ''),
            'buyer_name': body.get('buyer_name', ''),
            'buyer_tin': body.get('buyer_tin', ''),
            'subtotal': Decimal(str(body.get('subtotal', 0))),
            'vatable_sales': Decimal(str(body.get('vatable_sales', 0))),
            'vat_exempt_sales': Decimal(str(body.get('vat_exempt_sales', 0))),
            'zero_rated_sales': Decimal(str(body.get('zero_rated_sales', 0))),
            'vat_amount': Decimal(str(body.get('vat_amount', 0))),
            'total': Decimal(str(body.get('total', 0))),
            'department': body.get('department', ''),
            'employee_name': body.get('employee_name', ''),
        }
    )

    return JsonResponse({
        'receipt_id': receipt.id,
        'created': created,
        'drive_file_id': receipt.drive_file_id,
    }, status=201 if created else 200)


@require_GET
def list_receipts(request):
    if _is_n8n_request(request):
        receipts = Receipt.objects.all()
    elif request.user and request.user.is_authenticated:
        receipts = get_user_receipts(request.user)
    else:
        return JsonResponse({'error': 'Not authenticated'}, status=401)

    category = request.GET.get('category')
    status = request.GET.get('status')
    start, end = parse_date_range(request)

    if category:
        receipts = receipts.filter(expense_category=category)
    if status:
        receipts = receipts.filter(status=status)
    if request.GET.get('start') or request.GET.get('end'):
        receipts = receipts.filter(expense_date__range=[start, end])

    data = receipts.values(
        'id', 'drive_file_id', 'drive_file_name', 'business_name',
        'document_type', 'expense_category', 'vat_type', 'status',
        'expense_date', 'total', 'vat_amount', 'department',
        'employee_name', 'receipt_number', 'tin', 'created_at',
        'bir_permit_number',
    )

    return JsonResponse({
        'receipts': [
            {
                **r,
                'expense_date': r['expense_date'].isoformat() if r['expense_date'] else None,
                'created_at': r['created_at'].isoformat(),
                'total': str(r['total']),
                'vat_amount': str(r['vat_amount']),
            }
            for r in data
        ],
        'total_count': receipts.count(),
    })


@require_GET
@require_auth
def get_receipt(request, receipt_id):
    try:
        receipt = Receipt.objects.get(
            Q(id=receipt_id) & (Q(user=request.user) | Q(user__isnull=True))
        )
    except Receipt.DoesNotExist:
        return JsonResponse({'error': 'Receipt not found'}, status=404)

    return JsonResponse({
        'id': receipt.id,
        'drive_file_id': receipt.drive_file_id,
        'drive_file_name': receipt.drive_file_name,
        'drive_folder_id': receipt.drive_folder_id,
        'drive_folder_name': receipt.drive_folder_name,
        'status': receipt.status,
        'ocr_raw_text': receipt.ocr_raw_text,
        'ocr_processed_at': receipt.ocr_processed_at.isoformat() if receipt.ocr_processed_at else None,
        'document_type': receipt.document_type,
        'vat_type': receipt.vat_type,
        'expense_category': receipt.expense_category,
        'business_name': receipt.business_name,
        'business_address': receipt.business_address,
        'tin': receipt.tin,
        'receipt_number': receipt.receipt_number,
        'bir_permit_number': receipt.bir_permit_number,
        'expense_date': receipt.expense_date.isoformat() if receipt.expense_date else None,
        'description': receipt.description,
        'buyer_name': receipt.buyer_name,
        'buyer_tin': receipt.buyer_tin,
        'subtotal': str(receipt.subtotal),
        'vatable_sales': str(receipt.vatable_sales),
        'vat_exempt_sales': str(receipt.vat_exempt_sales),
        'zero_rated_sales': str(receipt.zero_rated_sales),
        'vat_amount': str(receipt.vat_amount),
        'total': str(receipt.total),
        'department': receipt.department,
        'employee_name': receipt.employee_name,
        'created_at': receipt.created_at.isoformat(),
        'updated_at': receipt.updated_at.isoformat(),
    })


@require_GET
def list_processed_file_ids(request):
    if not _is_n8n_request(request):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    file_ids = list(Receipt.objects.values_list('drive_file_id', flat=True))
    return JsonResponse({'processed_file_ids': file_ids, 'count': len(file_ids)})


# ─────────────────────────────────────────────
# ANALYTICS ENDPOINTS
# ─────────────────────────────────────────────

@require_GET
@require_auth
def analytics_summary(request):
    start, end = parse_date_range(request)
    user = request.user

    base_qs = get_user_receipts(user).filter(status='processed', expense_date__range=[start, end])

    duration = max((end - start).days, 1)
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=duration)
    prev_qs = get_user_receipts(user).filter(status='processed', expense_date__range=[prev_start, prev_end])

    current_total = base_qs.aggregate(total=Sum('total'))['total'] or Decimal('0')
    prev_total = prev_qs.aggregate(total=Sum('total'))['total'] or Decimal('0')
    change_pct = float((current_total - prev_total) / prev_total * 100) if prev_total > 0 else 0.0

    summary = base_qs.aggregate(
        total_spend=Sum('total'),
        total_vat=Sum('vat_amount'),
        transaction_count=Count('id'),
        avg_transaction=Avg('total'),
        largest_transaction=Max('total'),
        smallest_transaction=Min('total'),
    )

    return JsonResponse({
        'period': {'start': start.isoformat(), 'end': end.isoformat()},
        'total_spend': str(summary['total_spend'] or 0),
        'total_vat': str(summary['total_vat'] or 0),
        'transaction_count': summary['transaction_count'] or 0,
        'avg_transaction': str(summary['avg_transaction'] or 0),
        'largest_transaction': str(summary['largest_transaction'] or 0),
        'smallest_transaction': str(summary['smallest_transaction'] or 0),
        'vs_previous_period': {
            'previous_total': str(prev_total),
            'change_amount': str(current_total - prev_total),
            'change_pct': round(change_pct, 2),
        },
    })


@require_GET
@require_auth
def analytics_by_category(request):
    start, end = parse_date_range(request)
    base_qs = get_user_receipts(request.user).filter(status='processed', expense_date__range=[start, end])
    grand_total = base_qs.aggregate(total=Sum('total'))['total'] or Decimal('1')

    categories = (
        base_qs
        .values('expense_category')
        .annotate(total_spend=Sum('total'), transaction_count=Count('id'), avg_spend=Avg('total'))
        .order_by('-total_spend')
    )

    return JsonResponse({
        'period': {'start': start.isoformat(), 'end': end.isoformat()},
        'by_category': [
            {
                'category': cat['expense_category'],
                'total_spend': str(cat['total_spend']),
                'transaction_count': cat['transaction_count'],
                'avg_spend': str(cat['avg_spend']),
                'percentage': round(float(cat['total_spend'] / grand_total * 100), 2),
            }
            for cat in categories
        ],
    })


@require_GET
@require_auth
def analytics_trends(request):
    today = timezone.now().date()
    twelve_months_ago = today.replace(day=1) - timedelta(days=365)

    base_qs = get_user_receipts(request.user).filter(
        status='processed', expense_date__gte=twelve_months_ago,
    )

    monthly = (
        base_qs
        .annotate(month=TruncMonth('expense_date'))
        .values('month')
        .annotate(total_spend=Sum('total'), total_vat=Sum('vat_amount'),
                  transaction_count=Count('id'), avg_spend=Avg('total'))
        .order_by('month')
    )

    from django.db.models.functions import ExtractWeekDay
    by_weekday = (
        base_qs
        .annotate(weekday=ExtractWeekDay('expense_date'))
        .values('weekday')
        .annotate(total_spend=Sum('total'), count=Count('id'))
        .order_by('weekday')
    )

    weekday_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

    return JsonResponse({
        'monthly_trend': [
            {
                'month': row['month'].strftime('%Y-%m'),
                'total_spend': str(row['total_spend']),
                'total_vat': str(row['total_vat'] or 0),
                'transaction_count': row['transaction_count'],
                'avg_spend': str(row['avg_spend']),
            }
            for row in monthly
        ],
        'by_weekday': [
            {
                'weekday': weekday_names[row['weekday'] - 1],
                'total_spend': str(row['total_spend']),
                'transaction_count': row['count'],
            }
            for row in by_weekday
        ],
    })


# ─────────────────────────────────────────────
# N8N PROXY (kept for OCR workflow compatibility)
# ─────────────────────────────────────────────

@csrf_exempt
@require_POST
def n8n_analytics_proxy(request):
    if not _is_n8n_request(request):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        body = {}

    from django.contrib.auth import get_user_model
    User = get_user_model()
    user = User.objects.filter(id=body.get('user_id')).first()
    if not user:
        return JsonResponse({'error': 'User not found'}, status=404)

    today = timezone.now().date()
    start = today.replace(day=1)
    prev_end = start - timedelta(days=1)
    prev_start = prev_end.replace(day=1)

    base_qs = get_user_receipts(user).filter(status='processed', expense_date__range=[start, today])
    prev_total = get_user_receipts(user).filter(
        status='processed', expense_date__range=[prev_start, prev_end],
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')

    summary = base_qs.aggregate(
        total_spend=Sum('total'), total_vat=Sum('vat_amount'),
        transaction_count=Count('id'), avg_transaction=Avg('total'),
    )

    current_total = summary['total_spend'] or Decimal('0')
    change_pct = float((current_total - prev_total) / prev_total * 100) if prev_total > 0 else 0.0

    categories = list(base_qs.values('expense_category').annotate(total=Sum('total'), count=Count('id')).order_by('-total'))
    recent_receipts = list(
        get_user_receipts(user).filter(status='processed').order_by('-expense_date')[:20]
        .values('business_name', 'expense_category', 'document_type', 'total', 'vat_amount',
                'expense_date', 'description', 'tin', 'receipt_number', 'vat_type', 'bir_permit_number')
    )

    return JsonResponse({
        'summary': {
            'total_spend': str(current_total),
            'total_vat': str(summary['total_vat'] or 0),
            'transaction_count': summary['transaction_count'] or 0,
            'avg_transaction': str(summary['avg_transaction'] or 0),
            'period_start': start.isoformat(),
            'period_end': today.isoformat(),
            'vs_prev_month': {'previous_total': str(prev_total), 'change_pct': round(change_pct, 2)},
        },
        'by_category': [{**c, 'total': str(c['total'])} for c in categories],
        'recent_receipts': [
            {**r, 'expense_date': r['expense_date'].isoformat() if r['expense_date'] else None,
             'total': str(r['total']), 'vat_amount': str(r['vat_amount'])}
            for r in recent_receipts
        ],
    })


# ─────────────────────────────────────────────
# GOOGLE DRIVE HELPERS (for chat upload)
# ─────────────────────────────────────────────

def _get_drive_folders():
    """
    Returns all Google Drive folders as a list of dicts: [{id, name, parent_id}]
    Scoped to folders that contain 'lifewood' OR all folders if none found.
    """
    creds = _get_n8n_credentials()
    if not creds:
        return []
    from googleapiclient.discovery import build
    service = build('drive', 'v3', credentials=creds)
    results = service.files().list(
        q="mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name, parents)",
        pageSize=200,
        orderBy="name",
    ).execute()
    return results.get('files', [])


def _create_drive_folder(folder_name, parent_id=None):
    """Creates a new Google Drive folder and returns (id, name)."""
    creds = _get_n8n_credentials()
    if not creds:
        raise Exception('No Google Drive credentials available')
    from googleapiclient.discovery import build
    service = build('drive', 'v3', credentials=creds)
    metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
    }
    if parent_id:
        metadata['parents'] = [parent_id]
    folder = service.files().create(body=metadata, fields='id,name').execute()
    return folder['id'], folder['name']


def _upload_file_to_drive_folder(folder_id, file_obj, filename, mime_type):
    """Uploads a Django InMemoryUploadedFile to a Drive folder. Returns (file_id, file_name)."""
    import tempfile as tmpmod
    creds = _get_n8n_credentials()
    if not creds:
        raise Exception('No Google Drive credentials available')
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    service = build('drive', 'v3', credentials=creds)

    suffix = os.path.splitext(filename)[1] or '.jpg'
    with tmpmod.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        for chunk in file_obj.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        file_metadata = {'name': filename, 'parents': [folder_id]}
        media = MediaFileUpload(tmp_path, mimetype=mime_type or 'application/octet-stream', resumable=False)
        uploaded = service.files().create(body=file_metadata, media_body=media, fields='id,name').execute()
        return uploaded['id'], uploaded['name']
    finally:
        os.unlink(tmp_path)


# ─────────────────────────────────────────────
# SHARED OCR HELPER
# ─────────────────────────────────────────────

OCR_PROMPT = """You are a BIR (Bureau of Internal Revenue) receipt OCR specialist for the Philippines. Extract all structured data from this receipt or invoice image.

Return ONLY a valid JSON object — no markdown, no explanation, no code fences. Use these exact field names:
{
  "document_type": "official_receipt|invoice|sales_invoice|delivery_receipt|collection_receipt|acknowledgment_receipt|charge_invoice|cash_invoice|debit_memo|credit_memo|job_order|purchase_order|billing_statement|statement_of_account|unknown",
  "vat_type": "vat|non_vat|zero_rated|vat_exempt|unknown",
  "expense_category": "office_supplies|meals_entertainment|transportation|utilities|communication|professional_fees|rent|salaries|repairs_maintenance|taxes_licenses|insurance|advertising|miscellaneous|uncategorized",
  "business_name": "",
  "business_address": "",
  "tin": "",
  "receipt_number": "",
  "bir_permit_number": "",
  "expense_date": "YYYY-MM-DD or empty string",
  "description": "brief description of what was purchased",
  "buyer_name": "",
  "buyer_tin": "",
  "subtotal": 0.00,
  "vatable_sales": 0.00,
  "vat_exempt_sales": 0.00,
  "zero_rated_sales": 0.00,
  "vat_amount": 0.00,
  "total": 0.00
}

Rules:
- All numeric fields must be numbers (not strings)
- Dates must be YYYY-MM-DD format or empty string if not found
- TIN format: XXX-XXX-XXX-XXX
- If a field is not found on the receipt, use empty string or 0
- For expense_category, infer from the business name and description
- Return ONLY the JSON object, nothing else"""


def _run_ocr_and_save(file_id, file_name, folder_id, folder_name):
    """
    Downloads a Drive file, runs GPT-4o OCR via OpenRouter, saves Receipt.
    Shared by both the n8n endpoint and the chat upload endpoint.
    Returns the saved Receipt object.
    """
    creds = _get_n8n_credentials()
    if not creds:
        raise Exception('No stored Google credentials')

    from googleapiclient.discovery import build
    service = build('drive', 'v3', credentials=creds)
    metadata = service.files().get(fileId=file_id, fields="id,name,mimeType").execute()
    mime_type = metadata.get('mimeType', 'image/jpeg')
    content = service.files().get_media(fileId=file_id).execute()
    b64 = base64.b64encode(content).decode('utf-8')

    openrouter_key = os.environ.get('OPENROUTER_API_KEY')
    if not openrouter_key:
        raise Exception('OPENROUTER_API_KEY not configured')

    response = http_requests.post(
        'https://openrouter.ai/api/v1/chat/completions',
        headers={'Authorization': f'Bearer {openrouter_key}', 'Content-Type': 'application/json'},
        json={
            'model': 'openai/gpt-4o',
            'max_tokens': 1024,
            'messages': [{'role': 'user', 'content': [
                {'type': 'image_url', 'image_url': {'url': f'data:{mime_type};base64,{b64}', 'detail': 'high'}},
                {'type': 'text', 'text': OCR_PROMPT},
            ]}]
        },
        timeout=60,
    )
    response.raise_for_status()
    reply = response.json()['choices'][0]['message']['content']

    try:
        ocr_data = json.loads(reply.replace('```json', '').replace('```', '').strip())
    except Exception:
        ocr_data = {'document_type': 'unknown', 'vat_type': 'unknown',
                    'expense_category': 'uncategorized', 'business_name': '', 'total': 0}

    expense_date = None
    date_str = ocr_data.get('expense_date', '')
    if date_str:
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                expense_date = datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue

    receipt, _ = Receipt.objects.update_or_create(
        drive_file_id=file_id,
        defaults={
            'drive_file_name':    file_name,
            'drive_folder_id':    folder_id,
            'drive_folder_name':  folder_name,
            'status':             'processed',
            'ocr_raw_text':       reply,
            'ocr_processed_at':   timezone.now(),
            'document_type':      ocr_data.get('document_type',     'unknown'),
            'vat_type':           ocr_data.get('vat_type',          'unknown'),
            'expense_category':   ocr_data.get('expense_category',  'uncategorized'),
            'business_name':      ocr_data.get('business_name',     ''),
            'business_address':   ocr_data.get('business_address',  ''),
            'tin':                ocr_data.get('tin',               ''),
            'receipt_number':     ocr_data.get('receipt_number',    ''),
            'bir_permit_number':  ocr_data.get('bir_permit_number', ''),
            'expense_date':       expense_date,
            'description':        ocr_data.get('description',       ''),
            'buyer_name':         ocr_data.get('buyer_name',        ''),
            'buyer_tin':          ocr_data.get('buyer_tin',         ''),
            'subtotal':           Decimal(str(ocr_data.get('subtotal',         0) or 0)),
            'vatable_sales':      Decimal(str(ocr_data.get('vatable_sales',    0) or 0)),
            'vat_exempt_sales':   Decimal(str(ocr_data.get('vat_exempt_sales', 0) or 0)),
            'zero_rated_sales':   Decimal(str(ocr_data.get('zero_rated_sales', 0) or 0)),
            'vat_amount':         Decimal(str(ocr_data.get('vat_amount',       0) or 0)),
            'total':              Decimal(str(ocr_data.get('total',            0) or 0)),
        }
    )
    return receipt


# ─────────────────────────────────────────────
# OCR ENDPOINT (called by n8n Drive workflow)
# ─────────────────────────────────────────────

@csrf_exempt
def process_ocr(request):
    if not _is_n8n_request(request):
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON body'}, status=400)

    file_id     = data.get('file_id')
    folder_id   = data.get('folder_id', '')
    folder_name = data.get('folder_name', '')
    file_name   = data.get('file_name', '')

    if not file_id:
        return JsonResponse({'error': 'file_id is required'}, status=400)

    try:
        _run_ocr_and_save(file_id, file_name, folder_id, folder_name)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'status': 'ok', 'file_id': file_id, 'file_name': file_name})


# ─────────────────────────────────────────────
# CHAT RECEIPT UPLOAD (user uploads via chat)
# ─────────────────────────────────────────────

def _list_all_drive_folders(creds):
    """Return all non-trashed folders the user has access to."""
    from googleapiclient.discovery import build
    service = build('drive', 'v3', credentials=creds)
    results = service.files().list(
        q="mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id, name, parents)",
        pageSize=500,
        orderBy="name",
    ).execute()
    return results.get('files', [])


def _find_folder_by_name(folders, name):
    """Case-insensitive match: exact first, then partial."""
    name_lower = name.strip().lower()
    for f in folders:
        if f['name'].lower() == name_lower:
            return f
    for f in folders:
        if name_lower in f['name'].lower() or f['name'].lower() in name_lower:
            return f
    return None


def _create_drive_folder(creds, name, parent_id=None):
    """Create a new Google Drive folder, optionally inside a parent."""
    from googleapiclient.discovery import build
    service = build('drive', 'v3', credentials=creds)
    metadata = {'name': name, 'mimeType': 'application/vnd.google-apps.folder'}
    if parent_id:
        metadata['parents'] = [parent_id]
    return service.files().create(body=metadata, fields='id,name').execute()


def _upload_file_to_folder(creds, folder_id, file_content, file_name, mime_type):
    """Upload file bytes to a specific Drive folder."""
    import tempfile
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    service = build('drive', 'v3', credentials=creds)
    suffix = os.path.splitext(file_name)[1] or '.jpg'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(file_content)
        tmp_path = tmp.name
    try:
        metadata = {'name': file_name, 'parents': [folder_id]}
        media = MediaFileUpload(tmp_path, mimetype=mime_type, resumable=False)
        return service.files().create(
            body=metadata, media_body=media, fields='id,name,webViewLink'
        ).execute()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def _run_ocr_on_image(base64_image, mime_type, api_key):
    """Run BIR receipt OCR via OpenRouter and return parsed dict."""
    ocr_prompt = """You are a BIR receipt OCR specialist for the Philippines.
Extract all structured data from this receipt image.
Return ONLY a valid JSON object, no markdown, no explanation:
{
  "document_type": "official_receipt|invoice|sales_invoice|delivery_receipt|collection_receipt|acknowledgment_receipt|charge_invoice|cash_invoice|debit_memo|credit_memo|job_order|purchase_order|billing_statement|statement_of_account|unknown",
  "vat_type": "vat|non_vat|zero_rated|vat_exempt|unknown",
  "expense_category": "office_supplies|meals_entertainment|transportation|utilities|communication|professional_fees|rent|salaries|repairs_maintenance|taxes_licenses|insurance|advertising|miscellaneous|uncategorized",
  "business_name": "",
  "business_address": "",
  "tin": "",
  "receipt_number": "",
  "bir_permit_number": "",
  "expense_date": "YYYY-MM-DD or empty string",
  "description": "",
  "buyer_name": "",
  "buyer_tin": "",
  "subtotal": 0.00,
  "vatable_sales": 0.00,
  "vat_exempt_sales": 0.00,
  "zero_rated_sales": 0.00,
  "vat_amount": 0.00,
  "total": 0.00
}"""
    response = http_requests.post(
        'https://openrouter.ai/api/v1/chat/completions',
        headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
        json={
            'model': 'openai/gpt-4o',
            'max_tokens': 1024,
            'messages': [{'role': 'user', 'content': [
                {'type': 'image_url', 'image_url': {'url': f'data:{mime_type};base64,{base64_image}', 'detail': 'high'}},
                {'type': 'text', 'text': ocr_prompt},
            ]}],
        },
        timeout=60,
    )
    response.raise_for_status()
    raw = response.json()['choices'][0]['message']['content']
    return json.loads(raw.replace('```json', '').replace('```', '').strip())


@csrf_exempt
@require_auth
def upload_receipt_via_chat(request):
    """
    Called when the user uploads a receipt image in the chat panel.

    Flow:
    1. Read the uploaded image + user message
    2. List all Drive folders
    3. Ask GPT-4o to identify the target folder (or confirm creation)
    4. Find or create the folder in Drive
    5. Upload the image to Drive
    6. Run OCR and save the Receipt record
    7. Return a friendly AI reply summarising what happened
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded_file = request.FILES.get('file')
    user_message  = request.POST.get('message', '').strip() or 'Process this receipt.'
    conversation_id = request.POST.get('conversation_id') or None

    if not uploaded_file:
        return JsonResponse({'error': 'No file provided'}, status=400)

    # ── Drive credentials ──────────────────────────────────────────────────
    from google_drive.utils import get_user_drive_credentials
    creds = get_user_drive_credentials(request.user)
    if not creds:
        return JsonResponse({'error': 'Google Drive not connected. Please reconnect.'}, status=401)

    api_key = os.environ.get('OPENROUTER_API_KEY', '')
    if not api_key:
        return JsonResponse({'error': 'OPENROUTER_API_KEY not configured'}, status=500)

    # ── Get or create conversation ─────────────────────────────────────────
    if conversation_id:
        try:
            conversation = Conversation.objects.get(id=conversation_id, user=request.user)
        except Conversation.DoesNotExist:
            conversation = Conversation.objects.create(
                user=request.user, title=user_message[:60])
    else:
        conversation = Conversation.objects.create(
            user=request.user, title=user_message[:60])

    # Save user message (include filename for context)
    user_chat_msg = ChatMessage.objects.create(
        conversation=conversation,
        role='user',
        content=f'[Receipt upload: {uploaded_file.name}] {user_message}',
    )

    # ── Read file ──────────────────────────────────────────────────────────
    file_content  = uploaded_file.read()
    mime_type     = uploaded_file.content_type or 'image/jpeg'
    file_name     = uploaded_file.name
    base64_image  = base64.b64encode(file_content).decode('utf-8')

    # ── List all Drive folders ─────────────────────────────────────────────
    try:
        all_folders = _list_all_drive_folders(creds)
        folder_list_text = '\n'.join(
            f"  - {f['name']} (id: {f['id']})" for f in all_folders[:100]
        ) or '  (no folders found)'
    except Exception as e:
        print(f'Drive folder list error: {e}')
        all_folders = []
        folder_list_text = '(could not fetch folders)'

    # ── Ask GPT-4o to parse folder intent ─────────────────────────────────
    intent_prompt = f"""The user uploaded a receipt image and said: "{user_message}"

Available Google Drive folders:
{folder_list_text}

Determine where this receipt should be uploaded.

Return ONLY a JSON object:
{{
  "target_folder_name": "the folder name (from list or new name user wants)",
  "folder_id": "matching id from the list above if found, else null",
  "create_folder": true or false,
  "parent_folder_id": "id of parent folder if creating inside one, else null",
  "message": "one-line summary of what you will do"
}}

Rules:
- Match folder names case-insensitively, partial match is fine
- If the user says "create it" or the folder clearly doesn't exist, set create_folder=true
- If the folder exists in the list, set folder_id and create_folder=false
- Return ONLY the JSON, no markdown"""

    try:
        intent_resp = http_requests.post(
            'https://openrouter.ai/api/v1/chat/completions',
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            json={
                'model': 'openai/gpt-4o',
                'max_tokens': 300,
                'messages': [{'role': 'user', 'content': [
                    {'type': 'image_url', 'image_url': {
                        'url': f'data:{mime_type};base64,{base64_image}', 'detail': 'low'}},
                    {'type': 'text', 'text': intent_prompt},
                ]}],
            },
            timeout=30,
        )
        intent_resp.raise_for_status()
        intent_raw = intent_resp.json()['choices'][0]['message']['content']
        intent = json.loads(intent_raw.replace('```json', '').replace('```', '').strip())
    except Exception as e:
        print(f'Intent parse error: {e}')
        reply = ("I had trouble identifying which folder to use. "
                 "Please try again and mention the folder name clearly, e.g. "
                 "\"This receipt is for the Admin Expense folder.\"")
        agent_msg = ChatMessage.objects.create(
            conversation=conversation, role='agent', content=reply)
        conversation.save()
        return JsonResponse({
            'conversation_id': conversation.id,
            'reply': reply,
            'user_message_id': user_chat_msg.id,
            'agent_message_id': agent_msg.id,
        })

    # FIX: intent.get() default only fires when key MISSING, not when null.
    # If GPT returns "target_folder_name": null we get None, which then
    # propagates into _create_drive_folder causing "missing argument: name".
    folder_id   = intent.get('folder_id') or None
    folder_name = (intent.get('target_folder_name') or intent.get('folder_name') or '').strip() or None
    folder_created = False

    # If GPT couldn't determine a folder name, ask the user
    if not folder_name:
        available = ', '.join(f['name'] for f in all_folders[:10])
        reply = (
            "I couldn't determine which folder to use from your message. "
            "Please mention the folder name clearly, e.g. "
            "'This receipt is for the Admin Expense folder.' "
            f"Your existing folders: {available}"
        )
        agent_msg = ChatMessage.objects.create(
            conversation=conversation, role='agent', content=reply)
        conversation.save()
        return JsonResponse({
            'conversation_id': conversation.id, 'reply': reply,
            'user_message_id': user_chat_msg.id, 'agent_message_id': agent_msg.id,
        })

    # ── Find or create folder ──────────────────────────────────────────────
    if not folder_id:
        # One more attempt: fuzzy match locally before creating
        matched = _find_folder_by_name(all_folders, folder_name)
        if matched:
            folder_id   = matched['id']
            folder_name = matched['name']
        elif intent.get('create_folder'):
            try:
                parent_id  = intent.get('parent_folder_id') or None
                new_folder = _create_drive_folder(creds, folder_name, parent_id)
                folder_id  = new_folder['id']
                folder_created = True
            except Exception as e:
                reply = f"I couldn't create the folder **{folder_name}**: {e}"
                agent_msg = ChatMessage.objects.create(
                    conversation=conversation, role='agent', content=reply)
                conversation.save()
                return JsonResponse({
                    'conversation_id': conversation.id, 'reply': reply,
                    'user_message_id': user_chat_msg.id, 'agent_message_id': agent_msg.id,
                })
        else:
            # Folder not found and user didn't say to create — ask for clarification
            available = ', '.join(f['name'] for f in all_folders[:10])
            reply = (f"I couldn't find a folder named **{folder_name}**. "
                     f"Would you like me to create it? Just say \"yes, create it\" "
                     f"or choose from your existing folders:\n{available}")
            agent_msg = ChatMessage.objects.create(
                conversation=conversation, role='agent', content=reply)
            conversation.save()
            return JsonResponse({
                'conversation_id': conversation.id, 'reply': reply,
                'user_message_id': user_chat_msg.id, 'agent_message_id': agent_msg.id,
            })

    # ── Upload file to Drive ───────────────────────────────────────────────
    try:
        uploaded = _upload_file_to_folder(creds, folder_id, file_content, file_name, mime_type)
        drive_file_id = uploaded['id']
    except Exception as e:
        reply = f"I found the folder **{folder_name}** but couldn't upload the file: {e}"
        agent_msg = ChatMessage.objects.create(
            conversation=conversation, role='agent', content=reply)
        conversation.save()
        return JsonResponse({
            'conversation_id': conversation.id, 'reply': reply,
            'user_message_id': user_chat_msg.id, 'agent_message_id': agent_msg.id,
        })

    # ── OCR ────────────────────────────────────────────────────────────────
    try:
        ocr_data = _run_ocr_on_image(base64_image, mime_type, api_key)
    except Exception as e:
        print(f'OCR error: {e}')
        ocr_data = {
            'document_type': 'unknown', 'vat_type': 'unknown',
            'expense_category': 'uncategorized', 'total': 0,
        }

    # ── Parse expense date ─────────────────────────────────────────────────
    expense_date = None
    date_str = ocr_data.get('expense_date', '')
    if date_str:
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                expense_date = datetime.strptime(date_str, fmt).date()
                break
            except ValueError:
                continue

    # ── Save Receipt record ────────────────────────────────────────────────
    receipt, _ = Receipt.objects.update_or_create(
        drive_file_id=drive_file_id,
        defaults={
            'user':              request.user,
            'drive_file_name':   file_name,
            'drive_folder_id':   folder_id,
            'drive_folder_name': folder_name,
            'status':            'processed',
            'ocr_raw_text':      json.dumps(ocr_data),
            'ocr_processed_at':  timezone.now(),
            'document_type':     ocr_data.get('document_type',    'unknown'),
            'vat_type':          ocr_data.get('vat_type',         'unknown'),
            'expense_category':  ocr_data.get('expense_category', 'uncategorized'),
            'business_name':     ocr_data.get('business_name',    ''),
            'business_address':  ocr_data.get('business_address', ''),
            'tin':               ocr_data.get('tin',              ''),
            'receipt_number':    ocr_data.get('receipt_number',   ''),
            'bir_permit_number': ocr_data.get('bir_permit_number',''),
            'expense_date':      expense_date,
            'description':       ocr_data.get('description',      ''),
            'buyer_name':        ocr_data.get('buyer_name',       ''),
            'buyer_tin':         ocr_data.get('buyer_tin',        ''),
            'subtotal':          Decimal(str(ocr_data.get('subtotal',         0) or 0)),
            'vatable_sales':     Decimal(str(ocr_data.get('vatable_sales',    0) or 0)),
            'vat_exempt_sales':  Decimal(str(ocr_data.get('vat_exempt_sales', 0) or 0)),
            'zero_rated_sales':  Decimal(str(ocr_data.get('zero_rated_sales', 0) or 0)),
            'vat_amount':        Decimal(str(ocr_data.get('vat_amount',       0) or 0)),
            'total':             Decimal(str(ocr_data.get('total',            0) or 0)),
        }
    )

    # ── Build reply ────────────────────────────────────────────────────────
    action    = f"created **{folder_name}** and uploaded" if folder_created else f"uploaded to **{folder_name}**"
    business  = ocr_data.get('business_name') or 'Unknown'
    total_amt = float(ocr_data.get('total', 0) or 0)
    date_disp = ocr_data.get('expense_date') or 'Unknown date'
    doc_type  = ocr_data.get('document_type', 'unknown').replace('_', ' ').title()
    vat_type  = ocr_data.get('vat_type', 'unknown')

    reply = (
        f"Done! I've {action} your receipt.\n\n"
        f"**Receipt Summary:**\n"
        f"- Business: {business}\n"
        f"- Amount: PHP {total_amt:,.2f}\n"
        f"- Date: {date_disp}\n"
        f"- Document type: {doc_type}\n"
        f"- VAT type: {vat_type}"
    )

    # BIR compliance warnings
    warnings = []
    if not ocr_data.get('tin'):
        warnings.append("TIN is missing")
    if not ocr_data.get('receipt_number'):
        warnings.append("Receipt number is missing")
    if not ocr_data.get('bir_permit_number'):
        warnings.append("BIR permit number is missing")
    if warnings:
        reply += "\n\n⚠️ **BIR Compliance:** " + ", ".join(warnings) + "."

    agent_msg = ChatMessage.objects.create(
        conversation=conversation,
        role='agent',
        content=reply,
        metadata={'receipt_id': receipt.id, 'drive_file_id': drive_file_id, 'folder': folder_name},
    )
    conversation.save()

    return JsonResponse({
        'conversation_id': conversation.id,
        'reply':           reply,
        'user_message_id': user_chat_msg.id,
        'agent_message_id': agent_msg.id,
        'receipt_id':      receipt.id,
    })


# ─────────────────────────────────────────────
# CHAT RECEIPT UPLOAD
# ─────────────────────────────────────────────

@csrf_exempt
@require_auth
def upload_receipt_via_chat(request):
    """
    Accepts a receipt image + a natural language message via multipart form.
    - Parses which Drive folder the user wants using GPT
    - Creates the folder if it doesn't exist and user requests it
    - Uploads the file to Google Drive
    - Runs OCR and saves the receipt to the DB
    - Returns a conversational reply saved to chat history
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded_file = request.FILES.get('file')
    message = request.POST.get('message', '').strip()
    conversation_id = request.POST.get('conversation_id') or None

    if not uploaded_file:
        return JsonResponse({'error': 'No file provided'}, status=400)

    if not message:
        message = 'Please upload this receipt'

    # ── Get or create conversation ─────────────────────────────────────────
    if conversation_id:
        try:
            conversation = Conversation.objects.get(id=conversation_id, user=request.user)
        except Conversation.DoesNotExist:
            return JsonResponse({'error': 'Conversation not found'}, status=404)
    else:
        conversation = Conversation.objects.create(
            user=request.user,
            title=f'Receipt upload: {uploaded_file.name[:50]}',
        )

    # Save user message (with filename noted)
    ChatMessage.objects.create(
        conversation=conversation,
        role='user',
        content=f'{message} [Attached: {uploaded_file.name}]',
    )

    # ── Get available Drive folders ────────────────────────────────────────
    try:
        drive_folders = _get_drive_folders()  # [{id, name, parents}]
        folder_names = [f['name'] for f in drive_folders]
        folder_map = {f['name'].lower(): f for f in drive_folders}
    except Exception as e:
        print(f'Drive folder fetch error: {e}')
        drive_folders = []
        folder_names = []
        folder_map = {}

    # ── Use AI to parse folder intent ─────────────────────────────────────
    parse_prompt = f"""A user wants to upload a receipt to a specific Google Drive folder. Parse their message and determine the target folder.

User message: "{message}"

Available folders:
{json.dumps(folder_names, indent=2)}

Respond ONLY with a JSON object, no markdown, no explanation:
{{
  "target_folder_name": "the exact folder name from the list, or the new folder name the user wants to create",
  "matched_existing": true or false,
  "should_create": true or false,
  "confidence": "high|medium|low"
}}

Rules:
- Match case-insensitively and allow partial matches (e.g. "admin" matches "Admin Expense")
- Set matched_existing to true only if the folder name is in the available list
- Set should_create to true if the user says "create", "make", "there's no folder", "no folder yet", or similar
- If no folder is mentioned and confidence would be low, still return your best guess with confidence "low"
"""

    try:
        intent_reply, _ = _call_openrouter([{'role': 'user', 'content': parse_prompt}])
        intent = json.loads(intent_reply.replace('```json', '').replace('```', '').strip())
    except Exception as e:
        print(f'Intent parse error: {e}')
        intent = {'target_folder_name': None, 'matched_existing': False,
                  'should_create': False, 'confidence': 'low'}

    target_name = intent.get('target_folder_name', '')
    matched = intent.get('matched_existing', False)
    should_create = intent.get('should_create', False)
    confidence = intent.get('confidence', 'low')

    # ── Resolve folder ID ──────────────────────────────────────────────────
    folder_id = None
    resolved_folder_name = target_name
    action_log = ''

    if matched and target_name:
        # Find by case-insensitive match
        match = folder_map.get(target_name.lower())
        if not match:
            # Fuzzy fallback
            for key, val in folder_map.items():
                if target_name.lower() in key or key in target_name.lower():
                    match = val
                    break
        if match:
            folder_id = match['id']
            resolved_folder_name = match['name']
            action_log = f'uploaded to existing folder "{resolved_folder_name}"'

    if not folder_id and (should_create or not matched) and target_name:
        if should_create or confidence in ('high', 'medium'):
            try:
                folder_id, resolved_folder_name = _create_drive_folder(target_name)
                action_log = f'created new folder "{resolved_folder_name}" and uploaded receipt there'
            except Exception as e:
                reply = f"I couldn't create the folder \"{target_name}\": {str(e)}"
                ChatMessage.objects.create(conversation=conversation, role='agent', content=reply)
                conversation.save()
                return JsonResponse({'conversation_id': conversation.id, 'reply': reply, 'uploaded': False})

    # ── No folder resolved — ask for clarification ─────────────────────────
    if not folder_id:
        folder_list_text = '\n'.join(f'  • {n}' for n in folder_names[:25]) or '  (no folders found)'
        reply = (
            f"I need to know which folder to put this receipt in.\n\n"
            f"**Available folders:**\n{folder_list_text}\n\n"
            f"You can say something like:\n"
            f'  • *"This is for the Admin Expense folder"*\n'
            f'  • *"Put it in Condo Dues"*\n'
            f'  • *"Create a new folder called VIP Preparation and upload it there"*'
        )
        ChatMessage.objects.create(conversation=conversation, role='agent', content=reply)
        conversation.save()
        return JsonResponse({'conversation_id': conversation.id, 'reply': reply, 'uploaded': False})

    # ── Upload to Drive ────────────────────────────────────────────────────
    try:
        file_id, file_name = _upload_file_to_drive_folder(
            folder_id, uploaded_file, uploaded_file.name,
            uploaded_file.content_type or 'image/jpeg',
        )
    except Exception as e:
        reply = f"I found the folder \"{resolved_folder_name}\" but couldn't upload the file: {str(e)}"
        ChatMessage.objects.create(conversation=conversation, role='agent', content=reply)
        conversation.save()
        return JsonResponse({'conversation_id': conversation.id, 'reply': reply, 'uploaded': False})

    # ── Run OCR ────────────────────────────────────────────────────────────
    ocr_note = ''
    receipt_summary = ''
    try:
        receipt = _run_ocr_and_save(file_id, file_name, folder_id, resolved_folder_name)
        total = f'PHP {receipt.total:,.2f}' if receipt.total else 'amount not detected'
        merchant = receipt.business_name or 'merchant not detected'
        date = receipt.expense_date or 'date not detected'
        ocr_note = f'\n\n**OCR Result:**\n  • Merchant: {merchant}\n  • Amount: {total}\n  • Date: {date}'
    except Exception as e:
        print(f'OCR error for {file_id}: {e}')
        ocr_note = '\n\n*OCR processing will be picked up by the background workflow shortly.*'

    reply = (
        f"Receipt uploaded successfully!\n\n"
        f"**File:** {file_name}\n"
        f"**Folder:** {resolved_folder_name}"
        f"{ocr_note}"
    )

    ChatMessage.objects.create(
        conversation=conversation,
        role='agent',
        content=reply,
        metadata={
            'uploaded_file_id': file_id,
            'folder_id': folder_id,
            'folder_name': resolved_folder_name,
        },
    )
    conversation.save()

    return JsonResponse({
        'conversation_id': conversation.id,
        'reply': reply,
        'uploaded': True,
        'file_id': file_id,
        'folder_name': resolved_folder_name,
    })


# ─────────────────────────────────────────────
# EXCEL EXPORT ENDPOINT
# ─────────────────────────────────────────────

@require_GET
@require_auth
def export_receipts_excel(request):
    """
    Streams an Excel file of processed receipts.
    ?folder=<name>  — filter by drive_folder_name (icontains match)
    ?start=YYYY-MM-DD&end=YYYY-MM-DD  — optional date range filter
    Column headers have a green background as requested.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
    except ImportError:
        return JsonResponse({'error': 'openpyxl not installed. Add it to requirements.txt.'}, status=500)

    import io
    from django.utils import timezone as tz

    # ── Filters ────────────────────────────────────────────────────────────
    qs = get_user_receipts(request.user).filter(status='processed')

    folder_filter = request.GET.get('folder', '').strip()
    if folder_filter:
        qs = qs.filter(drive_folder_name__icontains=folder_filter)

    start_str = request.GET.get('start', '')
    end_str   = request.GET.get('end', '')
    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date   = datetime.strptime(end_str,   '%Y-%m-%d').date()
            qs = qs.filter(expense_date__range=[start_date, end_date])
        except ValueError:
            pass

    qs = qs.select_related('user').order_by('drive_folder_name', '-expense_date')
    receipts = list(qs)

    # ── Workbook ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Receipts Export'

    # ── Styles ─────────────────────────────────────────────────────────────
    # Green header background (Lifewood dark green)
    HEADER_FILL = PatternFill('solid', fgColor='046241')
    ALT_FILL    = PatternFill('solid', fgColor='F0FBF6')   # light mint alternating row
    WHITE_FILL  = PatternFill('solid', fgColor='FFFFFF')
    TOTAL_FILL  = PatternFill('solid', fgColor='FFB347')   # amber totals row

    HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    BODY_FONT   = Font(name='Calibri', size=10)
    BOLD_FONT   = Font(name='Calibri', bold=True, size=10)

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=False)
    RIGHT  = Alignment(horizontal='right',  vertical='center')

    thin = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    CURRENCY_FMT = '#,##0.00'
    DATE_FMT     = 'YYYY-MM-DD'
    DATETIME_FMT = 'YYYY-MM-DD HH:MM'

    # ── Title block ─────────────────────────────────────────────────────────
    num_cols = 24  # matches COLUMNS below
    end_col = get_column_letter(num_cols)

    ws.merge_cells(f'A1:{end_col}1')
    title_cell = ws['A1']
    title_cell.value     = 'Lifewood Expense AI — Receipts Export'
    title_cell.font      = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title_cell.fill      = HEADER_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{end_col}2')
    subtitle = ws['A2']
    subtitle.value = (
        f"Generated: {tz.now().strftime('%B %d, %Y %I:%M %p')}"
        + (f"  |  Folder: {folder_filter}" if folder_filter else "  |  All folders")
        + f"  |  {len(receipts)} records"
    )
    subtitle.font      = Font(name='Calibri', size=9, italic=True, color='333333')
    subtitle.alignment = CENTER
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 4   # spacer

    # ── Column definitions (exactly as requested) ──────────────────────────
    # (header_label, model_field_or_callable, col_width, fmt_type)
    COLUMNS = [
        ('User',              '__user__',           18, 'text'),
        ('File Name',         'drive_file_name',    30, 'text'),
        ('Folder ID',         'drive_folder_id',    28, 'text'),
        ('Folder Name',       'drive_folder_name',  26, 'text'),
        ('Status',            'status',             12, 'text'),
        ('OCR Processed At',  'ocr_processed_at',  20, 'datetime'),
        ('Document Type',     'document_type',      18, 'text'),
        ('VAT Type',          'vat_type',           14, 'text'),
        ('Expense Category',  'expense_category',   20, 'text'),
        ('Business Name',     'business_name',      28, 'text'),
        ('Business Address',  'business_address',   32, 'text'),
        ('TIN',               'tin',                16, 'text'),
        ('Receipt Number',    'receipt_number',     18, 'text'),
        ('BIR Permit No.',    'bir_permit_number',  18, 'text'),
        ('Expense Date',      'expense_date',       14, 'date'),
        ('Description',       'description',        34, 'text'),
        ('Buyer Name',        'buyer_name',         22, 'text'),
        ('Buyer TIN',         'buyer_tin',          16, 'text'),
        ('Subtotal',          'subtotal',           16, 'currency'),
        ('Vatable Sales',     'vatable_sales',      16, 'currency'),
        ('VAT Exempt Sales',  'vat_exempt_sales',   18, 'currency'),
        ('Zero Rated Sales',  'zero_rated_sales',   18, 'currency'),
        ('VAT Amount',        'vat_amount',         16, 'currency'),
        ('Total',             'total',              16, 'currency'),
    ]

    HEADER_ROW = 4

    # Write header row
    for col_idx, (label, _, width, _fmt) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────
    currency_col_indices = [
        i + 1 for i, (_, _, _, fmt) in enumerate(COLUMNS) if fmt == 'currency'
    ]

    for row_idx, receipt in enumerate(receipts, start=HEADER_ROW + 1):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, (_, field, _, fmt) in enumerate(COLUMNS, start=1):

            # Resolve value
            if field == '__user__':
                user_obj = receipt.user
                value = user_obj.get_full_name() or user_obj.email or user_obj.username if user_obj else '(unassigned)'
            else:
                value = getattr(receipt, field, None)

            # Type coerce
            if fmt == 'currency':
                value = float(value) if value is not None else 0.0
            elif fmt == 'datetime':
                if value and hasattr(value, 'replace'):
                    value = value.replace(tzinfo=None)
            elif fmt == 'date':
                pass   # keep as date; openpyxl handles it
            else:
                value = str(value) if value is not None else ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = BODY_FONT
            cell.fill   = fill
            cell.border = BORDER

            if fmt == 'currency':
                cell.number_format = CURRENCY_FMT
                cell.alignment     = RIGHT
            elif fmt in ('date', 'datetime'):
                cell.number_format = DATE_FMT if fmt == 'date' else DATETIME_FMT
                cell.alignment     = CENTER
            else:
                cell.alignment = LEFT

        ws.row_dimensions[row_idx].height = 16

    # ── Totals row ─────────────────────────────────────────────────────────
    if receipts:
        total_row = HEADER_ROW + len(receipts) + 1
        ws.row_dimensions[total_row].height = 20

        for col_idx in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=total_row, column=col_idx)
            c.fill   = TOTAL_FILL
            c.border = BORDER

        ws.cell(row=total_row, column=1, value='TOTALS').font      = BOLD_FONT
        ws.cell(row=total_row, column=1).alignment = CENTER

        for col_idx in currency_col_indices:
            start_r    = HEADER_ROW + 1
            end_r      = HEADER_ROW + len(receipts)
            col_letter = get_column_letter(col_idx)
            c = ws.cell(row=total_row, column=col_idx,
                        value=f'=SUM({col_letter}{start_r}:{col_letter}{end_r})')
            c.font          = BOLD_FONT
            c.number_format = CURRENCY_FMT
            c.alignment     = RIGHT

    # ── Freeze panes below header ──────────────────────────────────────────
    ws.freeze_panes = f'A{HEADER_ROW + 1}'

    # ── Build filename ─────────────────────────────────────────────────────
    timestamp = tz.now().strftime('%Y%m%d_%H%M')
    if folder_filter:
        safe_name = folder_filter.replace(' ', '_').replace('/', '-')[:40]
        filename  = f'lifewood_{safe_name}_{timestamp}.xlsx'
    else:
        filename = f'lifewood_receipts_{timestamp}.xlsx'

    # ── Stream ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    return response